"""
ShipScan v6  (speed-optimized)
  Tab 1 — Targeted Extract: filter by filename/pattern, auto-download matches
  Tab 2 — Browse & Pick: list ALL PDFs from sender newest→oldest, checkbox select, save chosen

WHAT CHANGED vs v5 (and why it's fast now)
  v5 downloaded every matching email IN FULL (RFC822) — including every PDF
  attachment — one email at a time, just to read attachment filenames. With
  2000+ emails that's hundreds of MB / GB pulled over the wire to find ~20 files.

  v6 instead:
    1. Asks the server for BODYSTRUCTURE + ENVELOPE only (tiny metadata, no
       attachment bytes), in BATCHES of a few hundred messages per round-trip.
    2. Reads attachment filenames straight from that metadata and matches them
       against your filters.
    3. Downloads the actual PDF bytes ONLY for the parts that matched (and that
       aren't already on disk), fetching just that one MIME part — not the email.

  Net effect: the "find the right files" phase moves almost entirely server-side,
  and you only transfer the PDFs you actually keep.

Dependencies:  flask  openpyxl  imapclient
"""

import os
import re
import io
import base64
import quopri
import fnmatch
import zipfile
import threading
from functools import wraps
from datetime import datetime, timedelta, date
from email.header import decode_header

from flask import Flask, render_template, request, jsonify, send_file, Response
from imapclient import IMAPClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── Cloud config ──────────────────────────────────────────────────────────────
# On Render the disk is ephemeral and there is no C:\ path, so all output goes to
# a server-side temp dir and is delivered to the browser as a ZIP.
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "/tmp/shipscan_out")
GATE_PW    = os.environ.get("SHIPSCAN_PASSWORD", "")  # set this on Render to lock the site
IS_CLOUD   = bool(os.environ.get("RENDER") or GATE_PW)


@app.before_request
def _password_gate():
    """If SHIPSCAN_PASSWORD is set, require HTTP Basic auth (any username + that password)."""
    if not GATE_PW:
        return  # no gate locally
    auth = request.authorization
    if not auth or auth.password != GATE_PW:
        return Response(
            "Authentication required.", 401,
            {"WWW-Authenticate": 'Basic realm="ShipScan"'},
        )


def _resolve_save_folder(client_value):
    """In the cloud, ignore the client's (Windows) path and use the server output dir."""
    return OUTPUT_DIR if IS_CLOUD else (client_value or "").strip()


IMAP_HOST = "imap.gmail.com"
BATCH = 300  # messages per metadata round-trip

# ── Job state (targeted extract) ─────────────────────────────────────────────
job = {"running": False, "done": False, "logs": [], "rows": [], "summary": {}, "excel_path": ""}

# ── Browse state ──────────────────────────────────────────────────────────────
browse = {"running": False, "done": False, "pdfs": [], "error": ""}


# ── Helpers ───────────────────────────────────────────────────────────────────

def decode_str(val):
    if not val:
        return ""
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8", errors="replace")
    out = []
    for part, cs in decode_header(val):
        if isinstance(part, bytes):
            out.append(part.decode(cs or "utf-8", errors="replace"))
        else:
            out.append(part)
    return "".join(out)


def make_server(gmail_addr, app_pass):
    server = IMAPClient(IMAP_HOST, ssl=True)
    server.login(gmail_addr, app_pass)
    return server


def matches_filter(filename, filters):
    a = filename.lower().strip()
    for f in filters:
        f = f.strip()
        if not f:
            continue
        fl = f.lower()
        if "*" in fl or "?" in fl:
            if fnmatch.fnmatch(a, fl):
                return True, f
        else:
            if a == fl or os.path.splitext(a)[0] == os.path.splitext(fl)[0]:
                return True, f
    return False, ""


def log(msg, level="info"):
    job["logs"].append({"t": datetime.now().strftime("%H:%M:%S"), "msg": msg, "lvl": level})


def _dec(x):
    return x.decode() if isinstance(x, (bytes, bytearray)) else (x or "")


def _walk_bodystructure(part, prefix=""):
    """
    Walk an imapclient BODYSTRUCTURE node WITHOUT downloading any bodies.
    Yields dicts: {part, mediatype, filename, enc_size, encoding} for every leaf part.
    'part' is the IMAP part-number string used to fetch that single part later.
    """
    out = []
    # A multipart node groups its children as a list/tuple at index 0.
    # A single-part node has the media-type string (bytes) at index 0.
    if isinstance(part[0], (list, tuple)):
        for i, child in enumerate(part[0], 1):
            num = f"{i}" if not prefix else f"{prefix}.{i}"
            out += _walk_bodystructure(child, num)
        return out

    num = prefix or "1"
    mediatype = (_dec(part[0]).lower() + "/" + _dec(part[1]).lower())
    encoding = _dec(part[5]).lower()
    enc_size = part[6] if len(part) > 6 and isinstance(part[6], int) else 0

    filename = None
    # Prefer Content-Disposition filename (index 8: (disp-type, (k, v, ...)))
    disp = part[8] if len(part) > 8 else None
    if isinstance(disp, (list, tuple)) and len(disp) > 1 and isinstance(disp[1], (list, tuple)):
        d = disp[1]
        for k in range(0, len(d) - 1, 2):
            if _dec(d[k]).lower() == "filename":
                filename = _dec(d[k + 1])
    # Fallback to Content-Type name param (index 2)
    if not filename and isinstance(part[2], (list, tuple)):
        ct = part[2]
        for k in range(0, len(ct) - 1, 2):
            if _dec(ct[k]).lower() == "name":
                filename = _dec(ct[k + 1])

    if filename:
        filename = decode_str(filename)

    out.append({
        "part": num,
        "mediatype": mediatype,
        "filename": filename,
        "enc_size": enc_size,
        "encoding": encoding,
    })
    return out


def _pdf_parts(bodystructure):
    """Return only the parts that are PDF attachments (by type or .pdf filename)."""
    res = []
    for p in _walk_bodystructure(bodystructure):
        fn = (p["filename"] or "")
        if p["mediatype"] == "application/pdf" or fn.lower().endswith(".pdf"):
            if fn:  # need a filename to match/save
                res.append(p)
    return res


def _decode_part_body(raw, encoding):
    """Decode a fetched MIME part body using its transfer encoding."""
    if raw is None:
        return b""
    if encoding == "base64":
        try:
            return base64.decodebytes(raw if isinstance(raw, bytes) else raw.encode())
        except Exception:
            return base64.b64decode(re.sub(rb"\s", b"", raw))
    if encoding in ("quoted-printable", "quopri"):
        return quopri.decodestring(raw if isinstance(raw, bytes) else raw.encode())
    return raw if isinstance(raw, bytes) else raw.encode()


def _decoded_kb(enc_size, encoding):
    real = enc_size * 0.75 if encoding == "base64" else enc_size
    return round(real / 1024, 1)


def _envelope_date(env):
    try:
        if env and env.date:
            return env.date.strftime("%Y-%m-%d %H:%M"), env.date.timestamp()
    except Exception:
        pass
    return "", 0.0


def _search_criteria(sender_email, days_back, subjects=None):
    """
    Build a server-side IMAP search. Adding SUBJECT terms makes Gmail return
    FEWER message IDs up front, so both the metadata scan and the download phase
    have less to do. SUBJECT is a case-insensitive substring match; multiple
    subject lines are OR'd together.
    """
    crit = ["FROM", sender_email]
    if days_back and days_back > 0:
        since = (datetime.now() - timedelta(days=days_back)).date()
        crit += ["SINCE", since]

    subs = [s.strip() for s in (subjects or []) if s and s.strip()]
    if len(subs) == 1:
        crit += ["SUBJECT", subs[0]]
    elif len(subs) > 1:
        # IMAP OR is binary, so nest: OR a (OR b c)
        node = ["SUBJECT", subs[-1]]
        for s in reversed(subs[:-1]):
            node = ["OR", ["SUBJECT", s], node]
        crit += node
    return crit


def _parse_subjects(config):
    raw = (config.get("subject", "") or "").strip()
    return [l.strip() for l in raw.splitlines() if l.strip()]


# ── MODE 1: Targeted extract ──────────────────────────────────────────────────

def run_targeted(config):
    job.update({"running": True, "done": False, "logs": [], "rows": [], "summary": {}})

    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    sender_email = config["sender_email"].strip()
    days_back    = int(config.get("days_back", 30))
    save_folder  = _resolve_save_folder(config.get("save_folder", ""))
    raw_filters  = config.get("filters", "").strip()
    filters      = [l.strip() for l in raw_filters.splitlines() if l.strip()]
    fetch_all    = len(filters) == 0

    excel_path = os.path.join(save_folder, f"shipscan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    job["excel_path"] = excel_path

    log(f"Mode: {'all PDFs' if fetch_all else str(len(filters)) + ' filter(s)'}")
    log("Connecting…")

    try:
        os.makedirs(save_folder, exist_ok=True)
        server = make_server(gmail_addr, app_pass)
        log("Connected", "success")
    except Exception as e:
        log(f"Login failed: {e}", "error")
        log("Enable IMAP and use an App Password.", "warn")
        job.update({"running": False, "done": True})
        return

    log(f"Searching from: {sender_email}")
    subjects = _parse_subjects(config)
    if subjects:
        log(f"Subject filter: {', '.join(subjects)}")
    try:
        server.select_folder("INBOX", readonly=True)
        uids = server.search(_search_criteria(sender_email, days_back, subjects))
        uids = sorted(uids, reverse=True)  # newest first
        log(f"{len(uids)} email(s) found", "success" if uids else "warn")
    except Exception as e:
        log(f"Search failed: {e}", "error")
        try: server.logout()
        except Exception: pass
        job.update({"running": False, "done": True})
        return

    # ── Phase 1: scan metadata only (no attachment bytes) ────────────────────
    # Collect matches as (uid, part-info, matched_filter, subject, email_date, sort_key)
    matches = []
    scanned = 0
    try:
        for start in range(0, len(uids), BATCH):
            chunk = uids[start:start + BATCH]
            meta = server.fetch(chunk, ["BODYSTRUCTURE", "ENVELOPE"])
            # preserve newest-first ordering within the chunk
            for uid in chunk:
                data = meta.get(uid)
                if not data:
                    continue
                bs = data.get(b"BODYSTRUCTURE")
                env = data.get(b"ENVELOPE")
                if bs is None:
                    continue
                subject = decode_str(env.subject) if (env and env.subject) else ""
                email_date, sort_key = _envelope_date(env)
                for p in _pdf_parts(bs):
                    fn = p["filename"]
                    if fetch_all:
                        matched_filter = "—"
                    else:
                        ok, matched_filter = matches_filter(fn, filters)
                        if not ok:
                            continue
                    matches.append({
                        "uid": uid, "part": p["part"], "encoding": p["encoding"],
                        "filename": fn, "matched_filter": matched_filter,
                        "subject": subject, "email_date": email_date, "sort_key": sort_key,
                    })
            scanned += len(chunk)
            log(f"Scanned {scanned}/{len(uids)} emails… ({len(matches)} match so far)")
    except Exception as e:
        log(f"Scan error: {e}", "error")

    log(f"{len(matches)} matching PDF(s) to download", "success" if matches else "warn")

    # ── Phase 2: download only matched parts (skip ones already on disk) ──────
    rows = []
    # group wanted parts per uid so each email is fetched at most once
    by_uid = {}
    for m in matches:
        by_uid.setdefault(m["uid"], []).append(m)

    for uid in sorted(by_uid.keys(), reverse=True):
        items = by_uid[uid]
        # figure out which of this email's parts still need downloading
        need = []
        for m in items:
            safe_fn = re.sub(r'[<>:"/\\|?*]', "_", m["filename"])
            m["save_path"] = os.path.join(save_folder, safe_fn)
            if os.path.exists(m["save_path"]):
                rows.append({"filename": m["filename"], "matched_filter": m["matched_filter"],
                             "email_date": m["email_date"], "subject": m["subject"],
                             "saved_path": m["save_path"], "status": "Already Exists"})
                job["rows"] = list(rows)
                log(f"⏭ Exists: {m['filename']}", "warn")
            else:
                need.append(m)
        if not need:
            continue

        try:
            keys = [f"BODY.PEEK[{m['part']}]" for m in need]
            resp = server.fetch([uid], keys)
            item = resp.get(uid, {})
            for m in need:
                # response key for BODY.PEEK[2] comes back as BODY[2]
                want = f"BODY[{m['part']}]".encode()
                raw = item.get(want)
                if raw is None:
                    for k, v in item.items():
                        if k.startswith(b"BODY[") and k.endswith(b"]") and v:
                            raw = v
                            break
                try:
                    pdf_data = _decode_part_body(raw, m["encoding"])
                    if not pdf_data:
                        raise ValueError("empty part")
                    with open(m["save_path"], "wb") as f:
                        f.write(pdf_data)
                    rows.append({"filename": m["filename"], "matched_filter": m["matched_filter"],
                                 "email_date": m["email_date"], "subject": m["subject"],
                                 "saved_path": m["save_path"], "status": "Saved"})
                    job["rows"] = list(rows)
                    log(f"✓ Saved: {m['filename']}", "success")
                except Exception as e:
                    rows.append({"filename": m["filename"], "matched_filter": m["matched_filter"],
                                 "email_date": m["email_date"], "subject": m["subject"],
                                 "saved_path": "", "status": "Failed"})
                    job["rows"] = list(rows)
                    log(f"✗ Failed: {m['filename']} — {e}", "error")
        except Exception as e:
            for m in need:
                rows.append({"filename": m["filename"], "matched_filter": m["matched_filter"],
                             "email_date": m["email_date"], "subject": m["subject"],
                             "saved_path": "", "status": "Failed"})
                job["rows"] = list(rows)
            log(f"✗ Fetch failed for one email — {e}", "error")

    try: server.logout()
    except Exception: pass

    if not rows:
        log("No matching PDFs found.", "warn")

    saved  = sum(1 for r in rows if r["status"] == "Saved")
    failed = sum(1 for r in rows if r["status"] == "Failed")
    exists = sum(1 for r in rows if r["status"] == "Already Exists")
    job["summary"] = {"total": len(rows), "saved": saved, "failed": failed, "exists": exists}

    if rows:
        try:
            _write_excel(rows, excel_path)
            log(f"Excel saved → {os.path.basename(excel_path)}", "success")
        except Exception as e:
            log(f"Excel error: {e}", "error")

    log(f"Done — {saved} saved · {failed} failed · {exists} existed",
        "success" if not failed else "warn")
    job.update({"running": False, "done": True})


def _write_excel(rows, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["#", "Filename", "Matched Filter", "Email Date", "Subject", "Saved Path", "Status"]
    hfill = PatternFill("solid", start_color="4F46E5")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[1].height = 26
    fills = {
        "Saved":          PatternFill("solid", start_color="DCFCE7"),
        "Failed":         PatternFill("solid", start_color="FEE2E2"),
        "Already Exists": PatternFill("solid", start_color="EDE9FE"),
    }
    dfont  = Font(name="Calibri", size=9)
    dalign = Alignment(vertical="center")
    for i, row in enumerate(rows):
        r = i + 2
        vals = [i + 1, row["filename"], row["matched_filter"], row["email_date"],
                row["subject"], row["saved_path"], row["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = dfont; cell.alignment = dalign; cell.border = thin
            if c == 7:
                cell.fill = fills.get(v, PatternFill())
    for c, w in enumerate([4, 30, 20, 16, 36, 46, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(path)


# ── MODE 2: Browse all PDFs ───────────────────────────────────────────────────

def run_browse(config):
    browse.update({"running": True, "done": False, "pdfs": [], "error": ""})

    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    sender_email = config["sender_email"].strip()
    days_back    = int(config.get("days_back", 30))

    try:
        server = make_server(gmail_addr, app_pass)
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Login failed: {e}"})
        return

    try:
        server.select_folder("INBOX", readonly=True)
        uids = server.search(_search_criteria(sender_email, days_back, _parse_subjects(config)))
        uids = sorted(uids, reverse=True)
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Search failed: {e}"})
        try: server.logout()
        except Exception: pass
        return

    pdfs = []
    try:
        for start in range(0, len(uids), BATCH):
            chunk = uids[start:start + BATCH]
            meta = server.fetch(chunk, ["BODYSTRUCTURE", "ENVELOPE"])
            for uid in chunk:
                data = meta.get(uid)
                if not data:
                    continue
                bs = data.get(b"BODYSTRUCTURE")
                env = data.get(b"ENVELOPE")
                if bs is None:
                    continue
                subject = decode_str(env.subject) if (env and env.subject) else ""
                email_date, sort_key = _envelope_date(env)
                for p in _pdf_parts(bs):
                    pdfs.append({
                        "id":         f"{uid}_{p['part']}",
                        "msg_id":     str(uid),
                        "uid":        uid,
                        "part":       p["part"],
                        "part_index": p["part"],
                        "encoding":   p["encoding"],
                        "filename":   p["filename"],
                        "email_date": email_date,
                        "sort_key":   sort_key,
                        "subject":    subject,
                        "size_kb":    _decoded_kb(p["enc_size"], p["encoding"]),
                    })
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Scan failed: {e}"})
        try: server.logout()
        except Exception: pass
        return

    try: server.logout()
    except Exception: pass

    pdfs.sort(key=lambda x: x["sort_key"], reverse=True)
    browse.update({"running": False, "done": True, "pdfs": pdfs, "error": ""})


def save_selected(config, selected_ids):
    """Download only the selected PDFs — fetching just their MIME parts."""
    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    save_folder  = _resolve_save_folder(config.get("save_folder", ""))

    os.makedirs(save_folder, exist_ok=True)

    # Group selected items by uid → list of parts to fetch.
    by_uid = {}
    for item in browse["pdfs"]:
        if item["id"] in selected_ids:
            by_uid.setdefault(item["uid"], []).append(item)

    try:
        server = make_server(gmail_addr, app_pass)
    except Exception as e:
        return {"ok": False, "error": str(e), "saved": [], "failed": []}

    server.select_folder("INBOX", readonly=True)
    saved_files, failed_files = [], []

    for uid, items in by_uid.items():
        try:
            keys = [f"BODY.PEEK[{it['part']}]" for it in items]
            resp = server.fetch([uid], keys)
            data = resp.get(uid, {})
            for it in items:
                raw = data.get(f"BODY[{it['part']}]".encode())
                if raw is None:
                    for k, v in data.items():
                        if k.startswith(b"BODY[") and k.endswith(b"]") and v:
                            raw = v
                            break
                try:
                    pdf_data = _decode_part_body(raw, it["encoding"])
                    if not pdf_data:
                        raise ValueError("empty part")
                    safe_fn   = re.sub(r'[<>:"/\\|?*]', "_", it["filename"])
                    save_path = os.path.join(save_folder, safe_fn)
                    with open(save_path, "wb") as f:
                        f.write(pdf_data)
                    saved_files.append(it["filename"])
                except Exception as e:
                    failed_files.append({"file": it["filename"], "error": str(e)})
        except Exception as e:
            for it in items:
                failed_files.append({"file": it["filename"], "error": str(e)})

    try: server.logout()
    except Exception: pass
    return {"ok": True, "saved": saved_files, "failed": failed_files}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# Mode 1
@app.route("/run", methods=["POST"])
def run():
    if job["running"]:
        return jsonify({"error": "Already running"}), 400
    threading.Thread(target=run_targeted, args=(request.json,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/status")
def status():
    return jsonify({
        "running": job["running"], "done": job["done"],
        "logs": job["logs"][-200:], "rows": job["rows"], "summary": job["summary"]
    })

@app.route("/download")
def download():
    p = job.get("excel_path", "")
    if not p or not os.path.exists(p):
        return "Run an extraction first.", 404
    return send_file(p, as_attachment=True, download_name=os.path.basename(p))


@app.route("/download_zip")
def download_zip():
    """Bundle every saved PDF + the Excel report into one ZIP for the browser."""
    folder = OUTPUT_DIR if IS_CLOUD else os.path.dirname(job.get("excel_path", "") or "")
    if not folder or not os.path.isdir(folder):
        return "Run an extraction first.", 404
    files = [f for f in os.listdir(folder)
             if f.lower().endswith((".pdf", ".xlsx"))
             and os.path.isfile(os.path.join(folder, f))]
    if not files:
        return "Nothing to download yet.", 404
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files:
            z.write(os.path.join(folder, f), arcname=f)
    buf.seek(0)
    name = f"shipscan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/zip")


# Mode 2
@app.route("/browse", methods=["POST"])
def browse_start():
    if browse["running"]:
        return jsonify({"error": "Browse already running"}), 400
    threading.Thread(target=run_browse, args=(request.json,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/browse_status")
def browse_status():
    return jsonify({
        "running": browse["running"], "done": browse["done"],
        "pdfs": browse["pdfs"], "error": browse["error"]
    })

@app.route("/save_selected", methods=["POST"])
def save_sel():
    data = request.json
    config       = data.get("config", {})
    selected_ids = set(data.get("selected", []))
    if not selected_ids:
        return jsonify({"error": "Nothing selected"}), 400
    result = save_selected(config, selected_ids)
    return jsonify(result)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n  ShipScan (cloud build)  →  http://0.0.0.0:{port}\n")
    app.run(host="0.0.0.0", port=port)
